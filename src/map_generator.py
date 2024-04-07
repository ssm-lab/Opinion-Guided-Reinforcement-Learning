import argparse
import gymnasium as gym
import imageio
import logging
import os
import pandas as pd
import random
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class MapTools():

    def __init__(self):
        self._FILES_PATH = 'src/files'
        self._HOLE_SEED = 100
        logging.basicConfig(format='[%(levelname)s] %(message)s')
        logging.getLogger().setLevel(logging.INFO)

    def index_to_cell(self, row, col):
        cell =  f'{get_column_letter(col+2)}{row+2}'
        logging.debug(cell)
        return cell

    def sequence_to_coordinates(self, sequence_number, edge):
        row = sequence_number // edge
        col = sequence_number % (edge)
        logging.debug(f'({row}, {col})')
        
        return (row, col)
        
    def randomize_holes(self, size, seed, frozen_tiles_ratio=0.8):
        num_holes = round(size*size*(1-frozen_tiles_ratio))-2
        logging.debug(num_holes)
        random.seed(seed)
        hole_cells = random.sample(range(1, size*size-1), num_holes) # encoded here that (0,0) is start and (29,29) is goal
        
        return hole_cells
        
    def apply_style(self, sheet, cell, symbol, font, color):
        sheet[cell] = symbol
        sheet[cell].font = font
        sheet[cell].fill = color
        
        thin = Side(border_style="thin", color="000000")
        sheet[cell].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def create_hole(self, sheet, cell):
        symbol = 'H'
        font = Font(color='FFFFFF', bold=True)
        color = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        
        self.apply_style(sheet, cell, symbol, font, color)
        
    def create_ice(self, sheet, cell):
        symbol = 'F'
        font = Font(color='000000')
        color = PatternFill(start_color='DAE9F8', end_color='DAE9F8', fill_type='solid')
        
        self.apply_style(sheet, cell, symbol, font, color)
        
    def create_start(self, sheet, cell):
        symbol = 'S'
        font = Font(color='000000', bold=True)
        color = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        self.apply_style(sheet, cell, symbol, font, color)
        
    def create_goal(self, sheet, cell):
        symbol = 'G'
        font = Font(color='000000', bold=True)
        color = PatternFill(start_color='83E28E', end_color='83E28E', fill_type='solid')
        
        self.apply_style(sheet, cell, symbol, font, color)

    def get_file_name(self, size, seed):
        return f'lake-{size}x{size}-seed{seed}'

    def generate_map(self, size, seed):
        file = os.path.abspath(f'{self._FILES_PATH}/{self.get_file_name(size, seed)}.xlsx')
        workbook = Workbook(file)
        
        workbook.save(filename=file)
        
        workbook = load_workbook(filename=file)
        sheet = workbook.active
        
        alignment = Alignment(horizontal='center', vertical='center')
        
        '''
        format
        '''
        for i in range(size+1):
            sheet.row_dimensions[i+1].height = 15
            sheet.column_dimensions[get_column_letter(i+1)].width = 2.75
            
            for j in range(size+1):
                cell = f'{get_column_letter(i+1)}{j+1}'
                sheet[cell].alignment = alignment
        
        '''
        column numbers
        '''
        for i in range(size):
            cell = f'{get_column_letter(i+2)}1'
            sheet[cell] = i
        
        '''
        row numbers
        '''        
        for i in range(size):
            cell = f'{get_column_letter(1)}{i+2}'
            sheet[cell] = i

        '''
        Create frozen tiles
        '''
        for i in range(1, size+1):
            for j in range(1, size+1):
                cell = f'{get_column_letter(i+1)}{j+1}'
                self.create_ice(sheet, cell)
        '''
        Create start and goal
        '''
        self.create_start(sheet, self.index_to_cell(0, 0))
        self.create_goal(sheet, self.index_to_cell(size-1, size-1))
        
        '''
        Create holes
        '''
        holes = self.randomize_holes(size=size, seed=seed)
        for hole in holes:
            (row, col) = self.sequence_to_coordinates(hole, size)
            self.create_hole(sheet, self.index_to_cell(row, col))
        
        workbook.save(filename=file)
        
    def parse_map(self, size, seed):
        file = os.path.abspath(f'{self._FILES_PATH}/{self.get_file_name(size, seed)}.xlsx')
        workbook = load_workbook(filename=file)
        
        sheet = workbook.active
        first_row = sheet[1]
        size = int(first_row[-1].value)+1
        
        map_desc = []
        for row in sheet[2: size+1]:
            map_desc.append(''.join([cell.value for cell in row[1:size+1]]))
        
        return map_desc
        
    def render_map(self, size, seed):
        map_desc = self.parse_map(size, seed)
        env = gym.make('FrozenLake-v1', desc=map_desc, render_mode='rgb_array')
        env.reset()
        img = env.render()
        imgfile = os.path.abspath(f'{self._FILES_PATH}/{self.get_file_name(size, seed)}.png')
        imageio.imwrite(imgfile, img)
        env.close()

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('-generate', action='store_true')
    parser.add_argument('-render', action='store_true')
    parser.add_argument('-size')
    parser.add_argument('-seed')

    options = parser.parse_args()
    
    assert options.size
    assert options.seed
    size = int(options.size)
    seed = int(options.seed)
    
    map_tools = MapTools()
    if(options.generate):
        map_tools.generate_map(size, seed)
    if(options.render):
        map_tools.render_map(size, seed)
    else:
        raise Exception('Exactly one of [-generate | -render] should be chosen.')