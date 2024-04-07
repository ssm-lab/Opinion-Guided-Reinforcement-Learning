import logging
import os
import pandas as pd
import random
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import gymnasium as gym
import time
import imageio

#Constants
FILES_PATH = 'src/files'


logging.basicConfig(format='[%(levelname)s] %(message)s')
logging.getLogger().setLevel(logging.INFO)

def parse_map(filename):
    file = os.path.abspath(f'{FILES_PATH}/{filename}.xlsx')
    workbook = load_workbook(filename=file)
    
    sheet = workbook.active
    first_row = sheet[1]
    size = int(first_row[-1].value)+1
    
    #for row in range(2, size+2):
    #    print(sheet[row])
    matrix = []
    for row in sheet[2: size+1]:
        matrix.append(''.join([cell.value for cell in row[1:size+1]]))
    
    return matrix
'''''''''''''''''''''''''''''''''''''''''''''
Main
'''''''''''''''''''''''''''''''''''''''''''''
SIZE = 4
SEED = 10
file_name = f'lake-{SIZE}x{SIZE}-seed{SEED}'
map_desc = parse_map(file_name)
env = gym.make('FrozenLake-v1', desc=map_desc, render_mode='rgb_array')
env.reset()
img = env.render()

imgfile = os.path.abspath(f'{FILES_PATH}/{file_name}.png')
imageio.imwrite(imgfile, img)
env.close()